"""
extract_config.py  —  xlsm から対応表 (config) を再生成するユーティリティ
==========================================================================

将来 Excel 側でマグネット/ダクトのブロック対応表が更新された場合に、
この 1 本を実行すれば config/ 以下の JSON/CSV を作り直せる。
これにより「Excelが正、Pythonはそれを読むだけ」という運用も可能。

使い方:
    python extract_config.py  path/to/SKEKB_Layout.xlsm

必要: openpyxl, oletools
    pip install openpyxl oletools
"""

import csv
import json
import re
import sys
import zipfile
from pathlib import Path

CONFIG_DIR = Path(__file__).parent / "config"


def extract_block_tables(xlsm_path: str):
    """*_MagBlock / *_DuctBlock シートを JSON に書き出す。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)

    for ring in ("HER", "LER"):
        # --- MagBlock: BN -> ブロック名 ---
        mag = {}
        ws = wb[f"{ring}_MagBlock"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                mag[str(row[0]).strip()] = str(row[1]).strip()
        _write_json(f"{ring.lower()}_mag_blocks.json", mag)
        print(f"{ring}_MagBlock  -> {len(mag)} 件")

        # --- DuctBlock: BN -> [ブロック名,...] ---
        duct = {}
        ws = wb[f"{ring}_DuctBlock"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                blocks = [str(b).strip() for b in row[1:9]
                          if b and str(b).strip().lower() not in ("none", "", "*")]
                if blocks:
                    duct[str(row[0]).strip()] = blocks
        _write_json(f"{ring.lower()}_duct_blocks.json", duct)
        print(f"{ring}_DuctBlock -> {len(duct)} 件")


def extract_ordered_ducts(xlsm_path: str):
    """
    Lattice シートのダクトブロック・セルの背景色から「オーダー済み」集合を作る。
    旧VBAは緑(VBA ColorIndex 4 = openpyxl indexed 11)のセルのブロックに
    "_Or" を付けて挿入していた。その緑セルのブロック名を集めて保存する。
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsm_path)   # 色を読むので data_only/read_only にしない

    def is_ordered(cell):
        f = cell.fill
        if not f or not f.patternType:
            return False
        fg = f.fgColor
        return fg.type == "indexed" and fg.indexed == 11   # 緑 = ColorIndex 4

    for ring in ("HER", "LER"):
        ws = wb[f"{ring}_Lattice"]
        ordered = set()
        for row in ws.iter_rows(min_row=5):
            for cell in row[12:20]:          # M列以降 = ダクトブロック名
                v = cell.value
                if v in (None, "") or not is_ordered(cell):
                    continue
                ordered.add(str(v).strip())
        _write_json(f"{ring.lower()}_duct_ordered.json", sorted(ordered))
        print(f"{ring} ordered ducts -> {len(ordered)} 件")


def extract_duct_by_element(xlsm_path: str):
    """
    Lattice シートから「要素名 -> ダクトブロック(複数)」の直接マップを作る。
    旧VBAの複雑な DuctBN 導出を再現する代わりに、実際に各要素へ割り当てられた
    ダクトブロックを要素名で直接引けるようにする (取りこぼし防止に最も確実)。
    ブロック名は _Or なしの素の名前で保存し、_Or 付与は実行時に別途行う。
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    for ring in ("HER", "LER"):
        ws = wb[f"{ring}_Lattice"]
        elem_map = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            name = str(row[0]).strip() if row[0] else ""
            if not name or name == "$$$":
                continue
            ducts = [str(c).strip() for c in row[12:20]
                     if c not in (None, "")
                     and str(c).strip().lower() not in ("none", "", "*")]
            if ducts:
                elem_map[name] = ducts
        _write_json(f"{ring.lower()}_duct_by_element.json", elem_map)
        print(f"{ring} duct-by-element -> {len(elem_map)} 要素")


def extract_overrides(xlsm_path: str):
    """VBA から要素名固有の上書き規則を CSV に書き出す。"""
    try:
        from oletools.olevba import VBA_Parser
    except ImportError:
        print("oletools 未インストールのため override 抽出をスキップ")
        return

    vba_text = ""
    vp = VBA_Parser(xlsm_path)
    for _, _, _, code in vp.extract_macros():
        vba_text += code + "\n"

    for ring, sub in (("HER", "HERInsertMagnetBlockName"),
                      ("LER", "LERInsertMagnetBlockName")):
        start = vba_text.find(f"Sub {sub}()")
        end = vba_text.find("\nEnd Sub", start)
        body = vba_text[start:end] if start >= 0 else ""
        rules = _parse_named_overrides(body)
        _write_csv(f"{ring.lower()}_mag_overrides.csv", rules)
        print(f"{ring} overrides -> {len(rules)} 件")


def _parse_named_overrides(sub_text: str):
    rules = []
    pat = re.compile(r'If Cells\(i, 1\) = "([^"]+)" Then(.*?)End If\s*\n',
                     re.DOTALL)
    for m in pat.finditer(sub_text):
        name, body = m.group(1), m.group(2)
        conds = re.findall(r'BN = "([^"]+)" Then\s*\n\s*BN = "([^"]+)"', body)
        if conds:
            for expect, result in conds:
                rules.append((name, expect, result))
        else:
            simple = re.findall(r'BN = "([^"]+)"', body)
            if len(simple) == 1:
                rules.append((name, "*", simple[0]))
    return rules


def _write_json(name, data):
    with open(CONFIG_DIR / name, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _write_csv(name, rules):
    with open(CONFIG_DIR / name, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["element", "expect_bn", "result_bn"])
        w.writerows(rules)



def extract_mag_by_element(xlsm_path: str):
    """
    Lattice シートから「要素名 -> マグネットブロック」の直接マップを作る。
    複雑な BN 導出で取りこぼす特殊磁石(QK*,ZV*,BW* 等)を要素名で直接引くため。
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    for ring in ("HER", "LER"):
        ws = wb[f"{ring}_Lattice"]
        emap = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            name = str(row[0]).strip() if row[0] else ""
            if not name or name == "$$$":
                continue
            mag = row[11]
            if mag not in (None, "") and str(mag).strip().lower() not in ("none", ""):
                emap[name] = str(mag).strip()
        _write_json(f"{ring.lower()}_mag_by_element.json", emap)
        print(f"{ring} mag-by-element -> {len(emap)} 要素")



def extract_component_and_ducttype(xlsm_path: str):
    """
    HER/LER の Component シートと Duct_Type シートを config に取り込む。
    Duct_Table 生成 (duct_table.py) の入力台帳。dispog には無い手維持データ。
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)

    def rows_from(sheet, hrow=2):
        ws = wb[sheet]
        data = list(ws.iter_rows(min_row=hrow, values_only=True))
        header = [("" if c is None else str(c).strip()) for c in data[0]]
        out = []
        for r in data[1:]:
            if all(c in (None, "") for c in r):
                continue
            d = {}
            for i, h in enumerate(header):
                if h == "":
                    continue
                v = r[i] if i < len(r) else None
                d[h] = "" if v is None else (v if isinstance(v, (int, float)) else str(v).strip())
            out.append(d)
        return out

    for ring in ("HER", "LER"):
        comp = rows_from(f"{ring}_Component")
        _write_json(f"{ring.lower()}_component.json", comp)
        dt = rows_from(f"{ring}_Duct_Type")
        dmap = {}
        for d in dt:
            nm = d.get("Duct Name", "")
            if nm:
                dmap[nm] = d
        _write_json(f"{ring.lower()}_ducttype.json", dmap)
        print(f"{ring} Component -> {len(comp)} 行 / Duct_Type -> {len(dmap)} 種類")



def extract_mag_others(xlsm_path: str):
    """各リングの Duct_Table から Mag Others 列を行位置で抽出 (手維持データ)。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    for ring in ("HER", "LER"):
        ws = wb[f"{ring}_Duct_Table"]
        vals = []
        for r in ws.iter_rows(min_row=3, values_only=True):
            if r[0] is None or str(r[0]).startswith("****"):
                break
            vals.append("" if r[4] is None else str(r[4]))
        _write_json(f"{ring.lower()}_mag_others.json", vals)
        print(f"{ring} Mag Others -> {len(vals)} 行")


def consolidate():
    """
    抽出した分割ファイルを統合形式にまとめる:
      {ring}_ducts.json   … 要素名 -> 最終ブロック名リスト ("_Or" 焼き込み済み)
      {ring}_magnets.json … {"by_element": ..., "by_bn": ...}
    分割ファイル (旧形式) は config/legacy/ に退避する。
    """
    import shutil
    legacy = CONFIG_DIR / "legacy"
    legacy.mkdir(exist_ok=True)
    for r in ("her", "ler"):
        be = json.load(open(CONFIG_DIR / f"{r}_duct_by_element.json", encoding="utf-8"))
        ordered = set(json.load(open(CONFIG_DIR / f"{r}_duct_ordered.json", encoding="utf-8")))
        ducts = {k: [b + "_Or" if b in ordered else b for b in v] for k, v in be.items()}
        _write_json(f"{r}_ducts.json", ducts)

        mbe = json.load(open(CONFIG_DIR / f"{r}_mag_by_element.json", encoding="utf-8"))
        mbn = json.load(open(CONFIG_DIR / f"{r}_mag_blocks.json", encoding="utf-8"))
        _write_json(f"{r}_magnets.json", {"by_element": mbe, "by_bn": mbn})

        for f in (f"{r}_duct_by_element.json", f"{r}_duct_blocks.json",
                  f"{r}_duct_ordered.json", f"{r}_mag_by_element.json",
                  f"{r}_mag_blocks.json", f"{r}_mag_overrides.csv"):
            p = CONFIG_DIR / f
            if p.exists():
                shutil.move(str(p), str(legacy / f))
        print(f"{r}: 統合 -> {r}_ducts.json ({len(ducts)}), "
              f"{r}_magnets.json (by_element {len(mbe)} / by_bn {len(mbn)})")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    CONFIG_DIR.mkdir(exist_ok=True)
    xlsm = sys.argv[1]
    print(f"対象: {xlsm}\n")
    extract_block_tables(xlsm)
    extract_ordered_ducts(xlsm)
    extract_duct_by_element(xlsm)
    extract_mag_by_element(xlsm)
    extract_component_and_ducttype(xlsm)
    extract_mag_others(xlsm)
    extract_overrides(xlsm)
    consolidate()
    print("\nconfig/ を更新しました (統合形式)。")
