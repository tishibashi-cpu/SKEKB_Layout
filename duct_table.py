"""
duct_table.py  —  HER_Duct_Table / LER_Duct_Table の再生成 (Component 駆動版)
============================================================================

旧VBA Make_HER_Duct_Table / Make_LER_Duct_Table が作っていた、リング全体の
マグネット配置・ダクト種類などをまとめた一覧表を再生成する。

【重要】Component シートは Duct_Table と「1行ずつ位置で対応」する背骨である
(HER 1105/1105, LER 1176/1176 行で (Mag名, Duct名) が完全一致することを確認済み)。
従って本ツールは Component を行の背骨として駆動し、

  * Component 由来の列 (Mag Type / BM Note / Q Support / Duct Name / IP,CCG / RP /
    VSW / GV / Bellows / Temp / Flow / 各設置日 / 真空 / NEG-L/R/Act / Duct length)
    … Component からそのまま転記 (xlsm と同一ソース・同一行順なので完全一致)
  * ダクト物理列 (Cross Section / Duct Note / BPM Height / NEG 群)
    … Duct_Type から Duct名で引く
  * 位置列 (Loc / S) … dispog から Mag名で引いて再計算 (IR 部は固定値)
  * Room … Mag名トリガーで D01..D12、さらに GV列のセクター標識でサブ表記
    (D01_IRL / D01_STP / …) を付与
  * Mag Others … dispog のステアリング系要素を直前磁石に集約 (近似。後述)

を組み合わせて構築する。

Mag Others について: 旧マクロは sler_1707 固有のハードコード例外で各ステアリングを
特定行へ細かく振り分けている。本ツールは「直前の磁石に集約」する近似で、名前は
正しい磁石付近に出るが、行単位の細かな配置や QKx 補正子の追加までは一致しない場合がある。
"""

from __future__ import annotations
import json
from pathlib import Path

import skekb_layout as sk

_CONFIG = Path(__file__).resolve().parent / "config"

# 出力列 (旧Duct_Table の 1..43 列目。44列目以降の集計は省略)
COLUMNS = [
    "Room", "Mag Name", "Loc.", "S [m]", "Mag Others", "Mag Type", "BM Note",
    "Q Support", "Duct Name", "Length [mm]", "Cross Section", "Duct Note",
    "BPM Height", "IP,CCG", "RP", "VSW", "GV", "NEG GP-50", "Other NEG",
    "NEG C200", "NEG Strip-L", "NEG Strip-R", "Bellows", "Temp-1", "Temp-2",
    "Temp-3", "Flow", "NEG Type-L", "NEG Trans-L", "NEG Pin-L", "NEG Type-R",
    "NEG Trans-R", "NEG Pin-R", "Duct Baked", "Duct Installed", "GV Installed",
    "Duct ID", "Bellows Installed", "Bellows ID", "Vacuum", "NEG-L", "NEG-R",
    "NEG Act.",
]

# Room 基本表記 D01..D12 の境界トリガー (この磁石名で以降を切替, VBA より)
_ROOM_TRIGGERS = {
    "HER": {"QD1E.2": "D12", "QD1E.5": "D11", "QDRNE.2": "D10", "QD1E.8": "D09",
            "QD1E.11": "D08", "QX4RE": "D07", "QD1E.14": "D06", "QD1E.17": "D05",
            "QFROE.2": "D04", "QD1E.20": "D03", "QD1E.23": "D02"},
    "LER": {"QD1P.2": "D12", "QD1P.5": "D11", "QFWNP.3": "D10", "QD1P.8": "D09",
            "QD1P.11": "D08", "QV3P.1": "D07", "QD1P.14": "D06", "QD1P.17": "D05",
            "QFWOP.3": "D04", "QD1P.20": "D03", "QD1P.23": "D02"},
}

# Loc(実験室からの距離)計算の octant 定数 (Duct_Table 版)
_LOC = {
    "HER": dict(nr=752.6, fr=1505.7, orr=2260.7, add=0.7),
    "LER": dict(nr=754.6, fr=1509.5, orr=2262.6, add=0.4),
}
_CIRC = 3016.315


def _loc_label(s: float, ring: str) -> str:
    c = _LOC[ring]
    isl = int(s / (_CIRC / 8.0))
    a = c["add"]
    if isl == 0: return "TL" + str(int(s + a))
    if isl == 1: return "NR" + str(int(c["nr"] - s - 0.2))
    if isl == 2: return "NL" + str(int(s + a - c["nr"]))
    if isl == 3: return "FR" + str(int(c["fr"] - s - 0.3))
    if isl == 4: return "FL" + str(int(s + 0.3 - c["fr"]))
    if isl == 5: return "OR" + str(int(c["orr"] - s - 0.2))
    if isl == 6: return "OL" + str(int(s + a - c["orr"]))
    if isl == 7: return "TR" + str(int(_CIRC - s - 0.2))
    return ""


# Step4: GV列(col17)の先頭6文字 -> GV$ (Room サブ表記)。初期 D01_IRL。
_GV_TRIGGERS = {
    "LER": {"D01_L1": "D01_STP", "D01_L2": "D01_WKB", "D01_L4": "D01_ARC",
            "D01_L5": "D12_ARC", "D11_L1": "D11_ARC", "D11_L2": "D11_WKB",
            "D11_L3": "D11_WIG", "D10_L1": "D10_WIG", "D10_L2": "D10_WKB",
            "D10_L3": "D10_ARC", "D10_L4": "D09_ARC", "D08_L1": "D08_ARC",
            "D08_L2": "D08_SRM", "D08_L4": "D08_ARS", "D08_L6": "D07_CRS",
            "D07_L1": "D07_ARS", "D07_L3": "D07_INJ", "D07_L4": "D07_ABT",
            "D07_L5": "D07_ARC", "D07_L6": "D06_TST", "D06_L1": "D06_ARC",
            "D05_L1": "D05_ARC", "D05_L2": "D05_WKB", "D05_L3": "D05_ARS",
            "D05_L4": "D05_WIG", "D04_L1": "D04_WIG", "D04_L2": "D04_WKB",
            "D04_L3": "D04_ARC", "D04_L4": "D03_ARC", "D02_L1": "D02_ARC",
            "D02_L2": "D02_WKB", "D02_L3": "D02_STP", "D02_L4": "D02_IRR"},
    "HER": {"D01_H1": "D01_STP", "D01_H2": "D01_WKB", "D01_H3": "D01_ARC",
            "D01_H4": "D12_ARC", "D12_H1": "D11_ARC", "D11_H1": "D11_WKB",
            "D11_H2": "D11_HOM", "D11_H3": "D11_SCC", "D11_H4": "D11_SCA",
            "D11_H5": "D10_STR", "D10_H1": "D10_SCA", "D10_H2": "D10_SCC",
            "D10_H3": "D10_HOM", "D10_H4": "D10_ARC", "D10_H5": "D09_ARC",
            "D08_H1": "D08_ARC", "D08_H2": "D08_INJ", "D08_H3": "D08_ABT",
            "D08_H4": "D07_CRS", "D07_H1": "D07_FBK", "D07_H2": "D07_MON",
            "D07_H3": "D07_ARC", "D07_H4": "D06_ARC", "D05_H1": "D05_ARC",
            "D05_H2": "D05_WKB", "D05_H3": "D05_WIG", "D05_H4": "D04_STR",
            "D04_H1": "D04_ARS", "D04_H3": "D04_SRM", "D04_H4": "D04_ARC",
            "D04_H5": "D03_ARC", "D02_H1": "D02_ARC", "D02_H2": "D02_WKB",
            "D02_H3": "D02_STP", "D02_H4": "D02_IRR"},
}

# IR cryostat 部の固定 Loc/S (dispog に無いため)
_IR_POS = {
    "LER": {"QC1LP": ("TL1", -0.97), "QC2LP": ("TL2", 3.21),
            "QC2RP": ("TR2", 3011.33), "QC1RP": ("TR0", "")},
    "HER": {"QC1LE": ("TL0", 2.51), "QC2LE": ("TL2", 2.51),
            "QC2RE": ("TR3", 3011.87), "QC1RE": ("TR1", "")},
}


def _load(name):
    return json.load(open(_CONFIG / name, encoding="utf-8"))


def _compute_mag_others(elements, by_element) -> dict:
    """dispog のステアリング系(Z/FZ/FQ/SD/SF/SL, ダクト無し)を直前磁石に集約。"""
    others: dict[str, list[str]] = {}
    last_mag = None
    for el in elements:
        base = el.name.lstrip("-")
        has_duct = bool(by_element.get(el.name) or by_element.get(base))
        if has_duct:
            last_mag = base
        elif (base[:1] == "Z" or base[:2] in ("FZ", "FQ", "SD", "SF", "SL")):
            if last_mag:
                others.setdefault(last_mag, []).append(base)
    return {m: "\n".join(v) for m, v in others.items()}


def _load_duct_map(ring: str) -> dict:
    """要素名 -> ダクト名リスト。統合 config (ducts.json) では挿入名に "_Or" が
    付くため、台帳 (Duct_Type) を引く用途向けに "_Or" を外して返す。"""
    try:
        d = _load(f"{ring.lower()}_ducts.json")
        return {k: [b[:-3] if b.endswith("_Or") else b for b in v]
                for k, v in d.items()}
    except FileNotFoundError:
        return _load(f"{ring.lower()}_duct_by_element.json")


def generate_duct_table(dispog_path: str, ring: str) -> list[dict]:
    """Component を背骨に Duct_Table の行(辞書)リストを生成する。"""
    elements = sk.parse_dispog(dispog_path)
    component = _load(f"{ring.lower()}_component.json")
    ducttype = _load(f"{ring.lower()}_ducttype.json")
    by_element = _load_duct_map(ring)

    # dispog: 基本磁石名 -> S (最初の出現)
    s_by_mag: dict[str, float] = {}
    for el in elements:
        s_by_mag.setdefault(el.name.lstrip("-"), el.s)

    others = _compute_mag_others(elements, by_element)
    ir = _IR_POS.get(ring, {})
    try:                                   # xlsm 抽出の Mag Others (行位置で適用)
        mag_others_cfg = _load(f"{ring.lower()}_mag_others.json")
    except Exception:
        mag_others_cfg = None

    rows: list[dict] = []
    for c in component:
        mag = str(c.get("Mag Name", "")).strip()
        duct = str(c.get("Duct Name", "")).strip()
        r = {col: "" for col in COLUMNS}

        # --- Component 由来 (xlsm と同一ソース) ---
        r["Mag Name"] = mag
        r["Mag Type"] = c.get("Mag Type", "")
        r["BM Note"] = c.get("BM Note", "")
        r["Q Support"] = c.get("Q Support", "")
        r["Duct Name"] = duct
        r["Length [mm]"] = c.get("Duct length", "")
        r["IP,CCG"] = c.get("IP, CCG", "")
        r["RP"] = c.get("RP", "")
        r["VSW"] = c.get("VSW", "")
        r["GV"] = c.get("GV", "")
        r["Bellows"] = c.get("Bellows-type", "")
        r["Temp-1"] = c.get("Temp-1", "")
        r["Temp-2"] = c.get("Temp-2", "")
        r["Temp-3"] = c.get("Temp-3", "")
        r["Flow"] = c.get("Vac. Flow No.", "")
        r["Duct Baked"] = c.get("ダクトベーキング", "")
        r["Duct Installed"] = c.get("ダクト設置日", "")
        r["GV Installed"] = c.get("GV設置日", "")
        r["Duct ID"] = c.get("ダクト名前", "")
        r["Bellows Installed"] = c.get("Bellows 設置日", "")
        r["Bellows ID"] = c.get("Bellows 名前", "")
        r["Vacuum"] = c.get("真空", "")
        r["NEG-L"] = c.get("NEG-L", "")
        r["NEG-R"] = c.get("NEG-R", "")
        r["NEG Act."] = c.get("NEG Act.", "")

        # --- ダクト物理列 (Duct_Type を Duct名で) ---
        dt = ducttype.get(duct, {})
        r["Cross Section"] = dt.get("Cross Sect", "") or dt.get("Cross Section", "")
        r["Duct Note"] = dt.get("Note", "") or c.get("Duct Note", "")
        r["BPM Height"] = dt.get("BPM Height", "")
        r["NEG GP-50"] = dt.get("NEG GP-50", "")
        r["Other NEG"] = dt.get("Other NEG", "")
        r["NEG C200"] = dt.get("NEG C200", "")
        r["NEG Strip-L"] = dt.get("NEG Strip-L", "")
        r["NEG Strip-R"] = dt.get("NEG Strip-R", "")
        r["NEG Type-L"] = dt.get("NEG Ty-L", "")
        r["NEG Trans-L"] = dt.get("Trans-L", "")
        r["NEG Pin-L"] = dt.get("Pin-L", "")
        r["NEG Type-R"] = dt.get("NEG Ty-R", "")
        r["NEG Trans-R"] = dt.get("Trans-R", "")
        r["NEG Pin-R"] = dt.get("Pin-R", "")
        cgv = c.get("GV", "")              # Component の GV 標識
        if not r["GV"]:                    # Component に標識が無ければ Duct_Type の d/u
            r["GV"] = dt.get("GV", "")
        # GV セクター境界行(標識あり・ダクト名なし)は Duct Note に "GV" を立てる
        # (旧マクロが col12="GV" を書き、その行の上下に罫線を引く挙動の再現)
        if cgv and not duct:
            r["Duct Note"] = "GV"

        # --- 位置列 (dispog / IR 固定) と Mag Others ---
        if mag:
            if mag in ir:
                r["Loc."], r["S [m]"] = ir[mag]
            else:
                s = s_by_mag.get(mag)
                if s is not None:
                    r["S [m]"] = int(s * 100.0) / 100.0
                    if mag[:1] in ("Q", "B"):
                        r["Loc."] = _loc_label(s, ring)
            if mag in others:
                r["Mag Others"] = others[mag]

        rows.append(r)

    # Mag Others は手維持データ。xlsm から抽出した config を行位置でそのまま適用し、
    # 完全一致させる (config が無い/行数が合わないときは上の近似計算を使う)。
    if mag_others_cfg and len(mag_others_cfg) == len(rows):
        for r, mo in zip(rows, mag_others_cfg):
            r["Mag Others"] = mo

    _assign_rooms(rows, ring)
    return rows


def _assign_rooms(rows, ring):
    """Room 基本表記 (D01..D12) と、GV列標識によるサブ表記を付与。"""
    triggers = _ROOM_TRIGGERS[ring]
    room = "D01"
    for r in rows:
        mag = r["Mag Name"]
        if mag in triggers:
            room = triggers[mag]
        r["Room"] = room

    gvtab = _GV_TRIGGERS.get(ring, {})
    gv = "D01_IRL"
    for r in rows:
        if r["Room"][:1] != "D":
            continue
        marker = str(r.get("GV", ""))[:6]
        if marker in gvtab:
            gv = gvtab[marker]
            continue                       # 切替行自体には付けない (gg=1 相当)
        r["Room"] = f"{r['Room']} ({gv})"


def write_duct_tables_xlsx(out_path: str,
                           her_dispog: str | None = None,
                           ler_dispog: str | None = None) -> dict:
    """
    HER_Duct_Table / LER_Duct_Table を持つ xlsx を書き出す。
    与えた dispog のリングだけシートを作る。戻り値: {sheet_name: 行数}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin")
    counts = {}

    for ring, dispog in (("HER", her_dispog), ("LER", ler_dispog)):
        if not dispog:
            continue
        rows = generate_duct_table(dispog, ring)
        ws = wb.create_sheet(f"{ring}_Duct_Table")
        ws.append([ring])
        ws.append(COLUMNS)
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.border = Border(bottom=thin)
        for r in rows:
            ws.append([r[c] for c in COLUMNS])

        # Duct Note がちょうど "GV"(GVセクター境界行)の行は、全列の上下に太め(thin)罫線。
        # ("GV Dummy" 等は通常行扱いなので対象外)
        ncols = len(COLUMNS)
        for ridx, r in enumerate(rows, start=3):       # データは3行目から
            if str(r.get("Duct Note", "")).strip() == "GV":
                for cc in range(1, ncols + 1):
                    cell = ws.cell(row=ridx, column=cc)
                    cur = cell.border
                    cell.border = Border(left=cur.left, right=cur.right,
                                         top=thin, bottom=thin)

        # 列幅
        for i, col in enumerate(COLUMNS, start=1):
            ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = \
                max(8, min(22, len(col) + 2))
        ws.freeze_panes = "A3"
        counts[f"{ring}_Duct_Table"] = len(rows)

    if not counts:
        raise SystemExit("HER か LER いずれかの dispog を指定してください。")
    wb.save(out_path)
    return counts


def _main(argv=None):
    import argparse
    p = argparse.ArgumentParser(
        description="HER_Duct_Table / LER_Duct_Table を xlsx に書き出す (Component駆動)")
    p.add_argument("--her", help="HER の dispog ファイル", default=None)
    p.add_argument("--ler", help="LER の dispog ファイル", default=None)
    p.add_argument("-o", "--out", default="Duct_Table.xlsx", help="出力 xlsx パス")
    a = p.parse_args(argv)
    if not a.her and not a.ler:
        p.error("--her か --ler のいずれか(または両方)を指定してください。")
    counts = write_duct_tables_xlsx(a.out, her_dispog=a.her, ler_dispog=a.ler)
    for sheet, n in counts.items():
        print(f"  {sheet}: {n} 行")
    print(f"出力: {a.out}")


if __name__ == "__main__":
    _main()
